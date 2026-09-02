#!/usr/bin/env python3
"""Inspect an XLSX workbook without modifying or saving it."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - depends on the caller's environment
    raise SystemExit(
        "ERROR: openpyxl is required. Install it with: python -m pip install openpyxl"
    ) from exc


ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def configure_utf8_output() -> None:
    """Prefer stable UTF-8 output across terminals."""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            reconfigure(encoding="utf-8", errors="replace")


def format_value(value: Any) -> str:
    if value is None or value == "":
        return "(none)"
    return str(value)


def limited(items: Iterable[str], max_items: int) -> tuple[list[str], int]:
    """Return a display-sized list and the number of omitted items."""
    values = list(items)
    if max_items == 0 or len(values) <= max_items:
        return values, 0
    return values[:max_items], len(values) - max_items


def print_list(
    label: str,
    items: Iterable[str],
    max_items: int,
    indent: str = "  ",
) -> None:
    values, omitted = limited(items, max_items)
    print(f"{label}:")
    if not values:
        print(f"{indent}(none)")
        return
    for value in values:
        print(f"{indent}- {value}")
    if omitted:
        print(f"{indent}- ... {omitted} more (use --max-items 0 to show all)")


def iter_materialized_cells(worksheet: Any) -> Iterable[Any]:
    """Iterate loaded cells without expanding a large sparse used range."""
    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict):
        yield from cells.values()
        return
    for row in worksheet.iter_rows():
        yield from row


def formula_text(cell: Any) -> str:
    value = cell.value
    if isinstance(value, str):
        return value
    return str(getattr(value, "text", value))


def iter_formula_cells(worksheet: Any) -> Iterable[Any]:
    for cell in iter_materialized_cells(worksheet):
        if cell.data_type == "f" or (
            isinstance(cell.value, str) and cell.value.startswith("=")
        ):
            yield cell


def iter_error_cells(worksheet: Any) -> Iterable[Any]:
    for cell in iter_materialized_cells(worksheet):
        value = cell.value
        if cell.data_type == "e" or (
            isinstance(value, str) and value.upper() in ERROR_VALUES
        ):
            yield cell


def error_cell_lines(worksheet: Any, cached_worksheet: Any) -> list[str]:
    lines: set[str] = set()
    for source, current_sheet in (
        ("stored cell", worksheet),
        ("cached value", cached_worksheet),
    ):
        for cell in iter_error_cells(current_sheet):
            lines.add(
                f"{cell.coordinate}: {format_value(cell.value)} "
                f"(source={source}, data_type={cell.data_type})"
            )
    return sorted(lines)


def defined_name_lines(workbook: Any) -> list[str]:
    lines: list[str] = []
    try:
        names = list(workbook.defined_names.values())
    except AttributeError:
        names = list(getattr(workbook.defined_names, "definedName", []))

    for item in names:
        name = getattr(item, "name", "(unnamed)")
        target = getattr(item, "attr_text", None) or getattr(item, "value", None)
        scope_id = getattr(item, "localSheetId", None)
        scope = "workbook"
        if isinstance(scope_id, int) and 0 <= scope_id < len(workbook.sheetnames):
            scope = workbook.sheetnames[scope_id]
        hidden = bool(getattr(item, "hidden", False))
        lines.append(
            f"{name} | scope={scope} | hidden={hidden} | target={format_value(target)}"
        )
    return lines


def custom_row_dimension_lines(worksheet: Any) -> list[str]:
    lines: list[str] = []
    for index, dimension in worksheet.row_dimensions.items():
        if (
            dimension.height is not None
            or dimension.hidden
            or dimension.outlineLevel
            or dimension.collapsed
        ):
            lines.append(
                f"row {index}: height={format_value(dimension.height)}, "
                f"hidden={bool(dimension.hidden)}, outline={dimension.outlineLevel}, "
                f"collapsed={bool(dimension.collapsed)}"
            )
    return lines


def custom_column_dimension_lines(worksheet: Any) -> list[str]:
    lines: list[str] = []
    for key, dimension in worksheet.column_dimensions.items():
        if (
            dimension.width is not None
            or dimension.hidden
            or dimension.outlineLevel
            or dimension.collapsed
        ):
            start = getattr(dimension, "min", None)
            end = getattr(dimension, "max", None)
            span = key
            if start and end and start != end:
                span = f"{get_column_letter(start)}:{get_column_letter(end)}"
            lines.append(
                f"column {span}: width={format_value(dimension.width)}, "
                f"hidden={bool(dimension.hidden)}, outline={dimension.outlineLevel}, "
                f"collapsed={bool(dimension.collapsed)}"
            )
    return lines


def table_lines(worksheet: Any) -> list[str]:
    lines: list[str] = []
    for name in worksheet.tables:
        table = worksheet.tables[name]
        style = getattr(getattr(table, "tableStyleInfo", None), "name", None)
        lines.append(f"{name}: range={table.ref}, style={format_value(style)}")
    return lines


def validation_lines(worksheet: Any) -> list[str]:
    lines: list[str] = []
    collection = getattr(worksheet, "data_validations", None)
    validations = getattr(collection, "dataValidation", []) if collection else []
    for validation in validations:
        parts = [
            f"ranges={format_value(validation.sqref)}",
            f"type={format_value(validation.type)}",
        ]
        if validation.operator:
            parts.append(f"operator={validation.operator}")
        if validation.formula1 is not None:
            parts.append(f"formula1={validation.formula1}")
        if validation.formula2 is not None:
            parts.append(f"formula2={validation.formula2}")
        parts.append(f"allow_blank={bool(validation.allow_blank)}")
        lines.append(" | ".join(parts))
    return lines


def conditional_formatting_lines(worksheet: Any) -> list[str]:
    lines: list[str] = []
    rules = getattr(worksheet.conditional_formatting, "_cf_rules", {})
    for target, target_rules in rules.items():
        sqref = getattr(target, "sqref", target)
        rule_types = [
            format_value(getattr(rule, "type", None)) for rule in target_rules
        ]
        lines.append(
            f"ranges={sqref} | rules={len(target_rules)} | "
            f"types={', '.join(rule_types)}"
        )
    return lines


def inspect_sheet(worksheet: Any, cached_worksheet: Any, max_items: int) -> None:
    formulas = list(iter_formula_cells(worksheet))
    merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]

    print()
    print(f"SHEET: {worksheet.title}")
    print(f"  State: {worksheet.sheet_state}")
    print(
        f"  Used range: {worksheet.calculate_dimension()} "
        f"(rows={worksheet.max_row}, columns={worksheet.max_column})"
    )
    print(f"  Freeze panes: {format_value(worksheet.freeze_panes)}")
    print(f"  AutoFilter: {format_value(worksheet.auto_filter.ref)}")
    print(f"  Protected: {bool(worksheet.protection.sheet)}")
    print(f"  Charts: {len(getattr(worksheet, '_charts', []))}")
    print(f"  Formula count: {len(formulas)}")
    print(f"  Data validation count: {len(validation_lines(worksheet))}")
    print(
        "  Conditional formatting range count: "
        f"{len(conditional_formatting_lines(worksheet))}"
    )

    print_list("  Merged ranges", merged_ranges, max_items, indent="    ")
    print_list("  Excel tables", table_lines(worksheet), max_items, indent="    ")
    print_list(
        "  Formulas",
        (f"{cell.coordinate}: {formula_text(cell)}" for cell in formulas),
        max_items,
        indent="    ",
    )
    print_list(
        "  Data validations",
        validation_lines(worksheet),
        max_items,
        indent="    ",
    )
    print_list(
        "  Conditional formatting",
        conditional_formatting_lines(worksheet),
        max_items,
        indent="    ",
    )
    print_list(
        "  Custom row dimensions",
        custom_row_dimension_lines(worksheet),
        max_items,
        indent="    ",
    )
    print_list(
        "  Custom column dimensions",
        custom_column_dimension_lines(worksheet),
        max_items,
        indent="    ",
    )
    print_list(
        "  Stored formula-error cells",
        error_cell_lines(worksheet, cached_worksheet),
        max_items,
        indent="    ",
    )


def inspect_workbook(path: Path, max_items: int) -> None:
    """Load and report workbook structure without saving it."""
    workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )
    cached_workbook = None
    try:
        cached_workbook = load_workbook(
            filename=path,
            read_only=False,
            data_only=True,
            keep_links=True,
        )
        print("XLSX WORKBOOK INSPECTION")
        print(f"Workbook filename: {path.name}")
        print(f"Workbook path: {path.resolve()}")
        print(f"Sheet count: {len(workbook.sheetnames)}")
        print_list("Sheet names", workbook.sheetnames, max_items)
        print_list("Defined names", defined_name_lines(workbook), max_items)
        for worksheet in workbook.worksheets:
            inspect_sheet(
                worksheet,
                cached_workbook[worksheet.title],
                max_items,
            )
        print()
        print("Read-only safety: workbook was opened for inspection and was not saved.")
        print(
            "Formula note: formulas are shown as stored; openpyxl did not calculate them."
        )
    finally:
        if cached_workbook is not None:
            cached_workbook.close()
        workbook.close()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inspect an XLSX workbook without modifying it."
    )
    parser.add_argument("xlsx_path", type=Path, help="Path to an existing .xlsx file")
    parser.add_argument(
        "--max-items",
        type=int,
        default=50,
        help="Maximum entries shown per detailed list; use 0 for unlimited (default: 50)",
    )
    args = parser.parse_args(argv)
    if args.max_items < 0:
        parser.error("--max-items must be zero or a positive integer")
    return args


def main(argv: Sequence[str] | None = None) -> int:
    configure_utf8_output()
    args = parse_args(argv)
    path: Path = args.xlsx_path

    if not path.exists():
        print(f"ERROR: File not found: {path}", file=sys.stderr)
        return 2
    if not path.is_file():
        print(f"ERROR: Path is not a file: {path}", file=sys.stderr)
        return 2
    if path.suffix.lower() != ".xlsx":
        print(f"ERROR: Expected an .xlsx file: {path}", file=sys.stderr)
        return 2

    try:
        inspect_workbook(path, args.max_items)
    except PermissionError:
        print(f"ERROR: Permission denied while reading: {path}", file=sys.stderr)
        return 2
    except (OSError, ValueError) as exc:
        print(f"ERROR: Could not inspect workbook: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # openpyxl exposes several parser-specific exceptions
        print(
            f"ERROR: Could not inspect workbook ({type(exc).__name__}): {exc}",
            file=sys.stderr,
        )
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
