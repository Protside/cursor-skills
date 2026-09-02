#!/usr/bin/env python3
"""Perform structural, formula, and feature checks on an XLSX workbook."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence
from zipfile import BadZipFile, ZipFile

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ImportError as exc:  # pragma: no cover - depends on the caller's environment
    raise SystemExit(
        "ERROR: openpyxl is required. Install it with: python -m pip install openpyxl"
    ) from exc


INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
FORMULA_ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?"}
MAX_EXCEL_ROW = 1_048_576
MAX_EXCEL_COLUMN = 16_384
DETAIL_LIMIT = 25


@dataclass(frozen=True)
class Finding:
    severity: str
    check: str
    message: str


def configure_utf8_output() -> None:
    """Prefer stable UTF-8 output across terminals."""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            reconfigure(encoding="utf-8", errors="replace")


def add_finding(
    findings: list[Finding], severity: str, check: str, message: str
) -> None:
    findings.append(Finding(severity, check, message))


def detail_summary(items: Iterable[str], limit: int = DETAIL_LIMIT) -> str:
    values = list(items)
    shown = values[:limit]
    suffix = ""
    if len(values) > limit:
        suffix = f"; ... {len(values) - limit} more"
    return "; ".join(shown) + suffix


def verify_zip_package(path: Path, findings: list[Finding]) -> bool:
    try:
        with ZipFile(path, "r") as archive:
            corrupt_member = archive.testzip()
    except BadZipFile as exc:
        add_finding(findings, "FAIL", "package", f"Not a valid XLSX ZIP package: {exc}")
        return False
    except OSError as exc:
        add_finding(findings, "FAIL", "package", f"Could not read XLSX package: {exc}")
        return False

    if corrupt_member:
        add_finding(
            findings,
            "FAIL",
            "package",
            f"Corrupt member in XLSX package: {corrupt_member}",
        )
        return False
    add_finding(findings, "PASS", "package", "XLSX ZIP package is readable.")
    return True


def validate_sheet_names(workbook: Any, findings: list[Finding]) -> None:
    names = workbook.sheetnames
    invalid: list[str] = []
    for name in names:
        reasons: list[str] = []
        if not name:
            reasons.append("empty")
        if len(name) > 31:
            reasons.append("longer than 31 characters")
        if INVALID_SHEET_CHARS.search(name):
            reasons.append("contains an invalid character")
        if name.startswith("'") or name.endswith("'"):
            reasons.append("begins or ends with an apostrophe")
        if reasons:
            invalid.append(f"{name!r} ({', '.join(reasons)})")

    folded = [name.casefold() for name in names]
    duplicates = sorted(
        {name for name in folded if folded.count(name) > 1}
    )

    if invalid:
        add_finding(
            findings,
            "FAIL",
            "sheet names",
            f"Invalid sheet names: {detail_summary(invalid)}",
        )
    else:
        add_finding(findings, "PASS", "sheet names", "All sheet names are valid.")

    if duplicates:
        add_finding(
            findings,
            "FAIL",
            "sheet names",
            "Sheet names are not unique (case-insensitive): "
            f"{detail_summary(duplicates)}",
        )
    else:
        add_finding(
            findings,
            "PASS",
            "sheet names",
            "Sheet names are unique (case-insensitive).",
        )


def iter_materialized_cells(worksheet: Any) -> Iterable[Any]:
    """Iterate loaded cells without expanding a large sparse used range."""
    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict):
        yield from cells.values()
        return
    for row in worksheet.iter_rows():
        yield from row


def iter_cells(workbook: Any) -> Iterable[tuple[Any, Any]]:
    for worksheet in workbook.worksheets:
        for cell in iter_materialized_cells(worksheet):
            yield worksheet, cell


def formula_text(cell: Any) -> str:
    value = cell.value
    if isinstance(value, str):
        return value
    return str(getattr(value, "text", value))


def validate_formulas(workbook: Any, findings: list[Finding]) -> None:
    formulas: list[tuple[str, str, str]] = []
    broken: list[str] = []

    for worksheet, cell in iter_cells(workbook):
        if cell.data_type == "f" or (
            isinstance(cell.value, str) and cell.value.startswith("=")
        ):
            formula = formula_text(cell)
            formulas.append((worksheet.title, cell.coordinate, formula))
            if "#REF!" in formula.upper():
                broken.append(f"{worksheet.title}!{cell.coordinate}: {formula}")

    if broken:
        add_finding(
            findings,
            "FAIL",
            "formulas",
            f"{len(broken)} formula(s) contain #REF!: {detail_summary(broken)}",
        )
    else:
        add_finding(
            findings,
            "PASS",
            "formulas",
            f"{len(formulas)} formula cell(s) found; none contain #REF!.",
        )

    broken_names: list[str] = []
    try:
        defined_names = list(workbook.defined_names.values())
    except AttributeError:
        defined_names = list(getattr(workbook.defined_names, "definedName", []))
    for defined_name in defined_names:
        target = getattr(defined_name, "attr_text", None)
        if isinstance(target, str) and "#REF!" in target.upper():
            broken_names.append(f"{defined_name.name}: {target}")
    if broken_names:
        add_finding(
            findings,
            "FAIL",
            "defined names",
            "Defined name(s) contain #REF!: "
            f"{detail_summary(broken_names)}",
        )


def collect_stored_errors(workbook: Any, source: str) -> set[tuple[str, str, str, str]]:
    errors: set[tuple[str, str, str, str]] = set()
    for worksheet, cell in iter_cells(workbook):
        value = cell.value
        normalized = value.upper() if isinstance(value, str) else None
        if normalized in FORMULA_ERROR_VALUES:
            errors.add((worksheet.title, cell.coordinate, normalized, source))
    return errors


def validate_stored_errors(
    formula_workbook: Any,
    cached_workbook: Any,
    findings: list[Finding],
) -> None:
    errors = collect_stored_errors(formula_workbook, "stored cell")
    errors.update(collect_stored_errors(cached_workbook, "cached value"))
    coordinates = sorted(
        {f"{sheet}!{cell}: {value} ({source})" for sheet, cell, value, source in errors}
    )
    if coordinates:
        add_finding(
            findings,
            "FAIL",
            "stored errors",
            f"{len(coordinates)} stored error value(s) found: "
            f"{detail_summary(coordinates)}",
        )
    else:
        add_finding(
            findings,
            "PASS",
            "stored errors",
            "No stored #REF!, #VALUE!, #DIV/0!, or #NAME? values found.",
        )


def normalize_sheet_token(token: str) -> str:
    token = token.strip()
    if len(token) >= 2 and token[0] == "'" and token[-1] == "'":
        return token[1:-1].replace("''", "'")
    return token


def validate_a1_range(
    range_text: str,
    worksheet_name: str,
    workbook: Any,
) -> tuple[bool, str]:
    """Validate a simple local or cross-sheet A1 reference."""
    expression = range_text.strip()
    if expression.startswith("="):
        expression = expression[1:].strip()
    if "[" in expression or "]" in expression:
        return False, "external workbook reference cannot be verified"

    target_sheet = worksheet_name
    cell_range = expression
    if "!" in expression:
        sheet_token, cell_range = expression.rsplit("!", 1)
        target_sheet = normalize_sheet_token(sheet_token)
        if target_sheet not in workbook.sheetnames:
            return False, f"references missing sheet {target_sheet!r}"

    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except (TypeError, ValueError):
        return False, "is not a simple A1 range"

    min_col = min_col or 1
    min_row = min_row or 1
    max_col = max_col or MAX_EXCEL_COLUMN
    max_row = max_row or MAX_EXCEL_ROW
    if (
        min_row < 1
        or min_col < 1
        or max_row > MAX_EXCEL_ROW
        or max_col > MAX_EXCEL_COLUMN
    ):
        return False, "falls outside Excel worksheet bounds"
    return True, f"valid range on {target_sheet}"


def workbook_defined_name_keys(workbook: Any) -> set[str]:
    try:
        names = {str(name).casefold() for name in workbook.defined_names.keys()}
        for worksheet in workbook.worksheets:
            local_names = getattr(worksheet, "defined_names", None)
            if local_names is not None:
                names.update(str(name).casefold() for name in local_names.keys())
        return names
    except AttributeError:
        return {
            str(getattr(item, "name", "")).casefold()
            for item in getattr(workbook.defined_names, "definedName", [])
        }


def validate_data_validations(workbook: Any, findings: list[Finding]) -> None:
    invalid: list[str] = []
    unverified: list[str] = []
    total = 0
    defined_names = workbook_defined_name_keys(workbook)

    for worksheet in workbook.worksheets:
        collection = getattr(worksheet, "data_validations", None)
        validations = getattr(collection, "dataValidation", []) if collection else []
        for index, validation in enumerate(validations, start=1):
            total += 1
            label = f"{worksheet.title} validation {index}"
            target_ranges = list(getattr(validation.sqref, "ranges", []))
            if not target_ranges:
                invalid.append(f"{label}: has no target range")
                continue

            for target in target_ranges:
                if (
                    target.min_row < 1
                    or target.min_col < 1
                    or target.max_row > MAX_EXCEL_ROW
                    or target.max_col > MAX_EXCEL_COLUMN
                ):
                    invalid.append(f"{label}: invalid target {target}")

            formula1 = validation.formula1
            if validation.type != "list" or formula1 in (None, ""):
                continue

            source = str(formula1).strip()
            if "#REF!" in source.upper():
                invalid.append(f"{label}: source contains #REF! ({source})")
                continue
            if source.startswith('"') and source.endswith('"'):
                continue

            candidate = source[1:] if source.startswith("=") else source
            valid_range, reason = validate_a1_range(source, worksheet.title, workbook)
            if valid_range:
                continue
            if candidate.casefold() in defined_names:
                continue
            if reason.startswith("references missing sheet"):
                invalid.append(f"{label}: {reason} ({source})")
            elif reason == "falls outside Excel worksheet bounds":
                invalid.append(f"{label}: {reason} ({source})")
            else:
                unverified.append(f"{label}: {source} ({reason})")

    if invalid:
        add_finding(
            findings,
            "FAIL",
            "data validation",
            f"{len(invalid)} invalid validation reference(s): "
            f"{detail_summary(invalid)}",
        )
    elif total:
        add_finding(
            findings,
            "PASS",
            "data validation",
            f"{total} data validation rule(s) have structurally valid target ranges.",
        )
    else:
        add_finding(
            findings,
            "PASS",
            "data validation",
            "No data validation rules are present.",
        )

    if unverified:
        add_finding(
            findings,
            "WARNING",
            "data validation",
            "Some list sources use dynamic, external, or non-simple expressions and "
            f"could not be verified statically: {detail_summary(unverified)}",
        )


def ranges_overlap(first: Any, second: Any) -> bool:
    return not (
        first.max_row < second.min_row
        or second.max_row < first.min_row
        or first.max_col < second.min_col
        or second.max_col < first.min_col
    )


def validate_merged_ranges(workbook: Any, findings: list[Finding]) -> None:
    invalid: list[str] = []
    overlaps: list[str] = []
    total = 0

    for worksheet in workbook.worksheets:
        ranges = list(worksheet.merged_cells.ranges)
        total += len(ranges)
        for merged in ranges:
            if (
                merged.min_row < 1
                or merged.min_col < 1
                or merged.max_row > MAX_EXCEL_ROW
                or merged.max_col > MAX_EXCEL_COLUMN
                or merged.min_row > merged.max_row
                or merged.min_col > merged.max_col
            ):
                invalid.append(f"{worksheet.title}!{merged}")

        ranges.sort(key=lambda item: (item.min_row, item.min_col))
        for index, first in enumerate(ranges):
            for second in ranges[index + 1 :]:
                if second.min_row > first.max_row:
                    break
                if ranges_overlap(first, second):
                    overlaps.append(f"{worksheet.title}: {first} overlaps {second}")

    if invalid or overlaps:
        messages: list[str] = []
        if invalid:
            messages.append(f"invalid ranges: {detail_summary(invalid)}")
        if overlaps:
            messages.append(f"overlapping ranges: {detail_summary(overlaps)}")
        add_finding(
            findings,
            "FAIL",
            "merged ranges",
            "; ".join(messages),
        )
    else:
        add_finding(
            findings,
            "PASS",
            "merged ranges",
            f"{total} merged range(s) are structurally valid and non-overlapping.",
        )


def detect_possible_formula_replacements(workbook: Any) -> list[str]:
    """Find non-formula cells directly between formulas in the same column."""
    candidates: list[str] = []
    for worksheet in workbook.worksheets:
        cells_by_column: dict[int, list[Any]] = {}
        for cell in iter_materialized_cells(worksheet):
            cells_by_column.setdefault(cell.column, []).append(cell)
        for column_cells in cells_by_column.values():
            formula_rows = {
                cell.row for cell in column_cells if cell.data_type == "f"
            }
            if len(formula_rows) < 2:
                continue
            for cell in column_cells:
                if (
                    cell.value not in (None, "")
                    and cell.data_type != "f"
                    and cell.row - 1 in formula_rows
                    and cell.row + 1 in formula_rows
                ):
                    candidates.append(
                        f"{worksheet.title}!{cell.coordinate}={cell.value!r}"
                    )
    return candidates


def validate_formula_continuity(workbook: Any, findings: list[Finding]) -> None:
    candidates = detect_possible_formula_replacements(workbook)
    if candidates:
        add_finding(
            findings,
            "WARNING",
            "formula continuity",
            "Possible static replacement(s) directly between formulas: "
            f"{detail_summary(candidates)}",
        )
    else:
        add_finding(
            findings,
            "PASS",
            "formula continuity",
            "No obvious static replacements between adjacent formulas were detected.",
        )
    add_finding(
        findings,
        "INFO",
        "formula continuity",
        "A single workbook cannot prove that every intended formula remains a formula; "
        "the check is heuristic and a known-good baseline would be required for certainty.",
    )


def validate_basic_structure(workbook: Any, findings: list[Finding]) -> None:
    if not workbook.sheetnames:
        add_finding(findings, "FAIL", "workbook structure", "Workbook has no sheets.")
        return

    add_finding(
        findings,
        "PASS",
        "workbook structure",
        f"Workbook contains {len(workbook.sheetnames)} sheet(s).",
    )

    invalid_dimensions: list[str] = []
    invalid_tables: list[str] = []
    for worksheet in workbook.worksheets:
        if (
            worksheet.max_row < 1
            or worksheet.max_column < 1
            or worksheet.max_row > MAX_EXCEL_ROW
            or worksheet.max_column > MAX_EXCEL_COLUMN
        ):
            invalid_dimensions.append(
                f"{worksheet.title}: rows={worksheet.max_row}, "
                f"columns={worksheet.max_column}"
            )
        for name in worksheet.tables:
            table = worksheet.tables[name]
            valid, reason = validate_a1_range(table.ref, worksheet.title, workbook)
            if not valid:
                invalid_tables.append(
                    f"{worksheet.title}/{name}: {table.ref} ({reason})"
                )

    if invalid_dimensions:
        add_finding(
            findings,
            "FAIL",
            "workbook structure",
            f"Invalid worksheet dimensions: {detail_summary(invalid_dimensions)}",
        )
    if invalid_tables:
        add_finding(
            findings,
            "FAIL",
            "workbook structure",
            f"Invalid Excel table ranges: {detail_summary(invalid_tables)}",
        )
    if not invalid_dimensions and not invalid_tables:
        add_finding(
            findings,
            "PASS",
            "workbook structure",
            "Worksheet dimensions and Excel table ranges are structurally valid.",
        )


def overall_status(findings: Sequence[Finding]) -> str:
    if any(item.severity == "FAIL" for item in findings):
        return "FAIL"
    if any(item.severity == "WARNING" for item in findings):
        return "WARNING"
    return "PASS"


def print_report(path: Path, findings: Sequence[Finding]) -> str:
    status = overall_status(findings)
    print("XLSX VALIDATION")
    print(f"Workbook: {path.name}")
    print(f"Path: {path.resolve()}")
    print(f"Overall: {status}")
    print()
    print("Findings:")
    for finding in findings:
        print(f"  [{finding.severity}] {finding.check}: {finding.message}")
    print()
    print(
        "Formula recalculation limitation: openpyxl does not calculate Excel "
        "formulas. This validator inspected stored formulas and cached values only; "
        "it did not recalculate the workbook and does not claim recalculated results."
    )
    return status


def validate_workbook(path: Path) -> tuple[str, list[Finding]]:
    findings: list[Finding] = []
    if not verify_zip_package(path, findings):
        return "FAIL", findings

    formula_workbook = None
    cached_workbook = None
    try:
        formula_workbook = load_workbook(
            filename=path,
            read_only=False,
            data_only=False,
            keep_links=True,
        )
        add_finding(
            findings,
            "PASS",
            "open",
            "Workbook opened successfully with formulas preserved.",
        )
        cached_workbook = load_workbook(
            filename=path,
            read_only=False,
            data_only=True,
            keep_links=True,
        )

        validate_basic_structure(formula_workbook, findings)
        validate_sheet_names(formula_workbook, findings)
        validate_formulas(formula_workbook, findings)
        validate_stored_errors(formula_workbook, cached_workbook, findings)
        validate_data_validations(formula_workbook, findings)
        validate_merged_ranges(formula_workbook, findings)
        validate_formula_continuity(formula_workbook, findings)
    except PermissionError:
        add_finding(findings, "FAIL", "open", f"Permission denied while reading {path}.")
    except (OSError, ValueError) as exc:
        add_finding(findings, "FAIL", "open", f"Workbook could not be opened: {exc}")
    except Exception as exc:  # openpyxl exposes several parser-specific exceptions
        add_finding(
            findings,
            "FAIL",
            "open",
            f"Workbook could not be validated ({type(exc).__name__}): {exc}",
        )
    finally:
        if cached_workbook is not None:
            cached_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()

    return overall_status(findings), findings


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate the structure and stored content of an XLSX workbook."
    )
    parser.add_argument("xlsx_path", type=Path, help="Path to an existing .xlsx file")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    configure_utf8_output()
    args = parse_args(argv)
    path: Path = args.xlsx_path

    if not path.exists():
        findings = [Finding("FAIL", "input", f"File not found: {path}")]
        print_report(path, findings)
        return 2
    if not path.is_file():
        findings = [Finding("FAIL", "input", f"Path is not a file: {path}")]
        print_report(path, findings)
        return 2
    if path.suffix.lower() != ".xlsx":
        findings = [Finding("FAIL", "input", f"Expected an .xlsx file: {path}")]
        print_report(path, findings)
        return 2

    status, findings = validate_workbook(path)
    print_report(path, findings)
    return {"PASS": 0, "WARNING": 1, "FAIL": 2}[status]


if __name__ == "__main__":
    raise SystemExit(main())
